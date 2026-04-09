import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Parse CSV
with open('/root/.openclaw/media/inbound/00e7123a-df38-40ad-ad9e-17842a9c9ee4.csv', 'r') as f:
    lines = f.readlines()

header = [h.strip() for h in lines[0].split(',')]
data_rows = []
for line in lines[1:]:
    if line.strip():
        data_rows.append([c.strip() for c in line.split(',')])

total_row = data_rows[-1]
detail_rows = data_rows[:-1]

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "ACH Payments"

HEADER_FILL = PatternFill("solid", fgColor="1a6880")
ALT_FILL = PatternFill("solid", fgColor="f0f8fa")
WHITE_FILL = PatternFill("solid", fgColor="ffffff")

# Header row
for col_idx, col_name in enumerate(header, 1):
    cell = ws.cell(row=1, column=col_idx, value=col_name)
    cell.font = Font(bold=True, color="ffffff", name="Calibri", size=10)
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

ws.row_dimensions[1].height = 32

# Data rows
for row_idx, row in enumerate(detail_rows, 2):
    fill = ALT_FILL if row_idx % 2 == 0 else WHITE_FILL
    for col_idx, value in enumerate(row, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.font = Font(name="Calibri", size=10)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="left", vertical="center")
        if col_idx in (14, 16, 17) and value:
            try:
                cell.value = float(value)
                cell.number_format = '"$"#,##0.00'
                cell.alignment = Alignment(horizontal="right")
            except:
                pass

# Totals row
for col_idx, value in enumerate(total_row, 1):
    cell = ws.cell(row=len(detail_rows)+2, column=col_idx, value=value)
    cell.font = Font(bold=True, name="Calibri", size=10, color="8B4513")
    cell.fill = PatternFill("solid", fgColor="FFE4B5")
    cell.alignment = Alignment(horizontal="left", vertical="center")
    if col_idx in (14, 16, 17) and value:
        try:
            cell.value = float(value)
            cell.number_format = '"$"#,##0.00'
            cell.alignment = Alignment(horizontal="right")
        except:
            pass

col_widths = [8, 36, 36, 40, 8, 12, 14, 16, 18, 16, 24, 28, 16, 10, 16, 16, 38, 14, 16, 14]
for i, width in enumerate(col_widths, 1):
    if i <= len(col_widths):
        ws.column_dimensions[get_column_letter(i)].width = width

ws.freeze_panes = "A2"
last_row = len(detail_rows) + 1
ws.auto_filter.ref = f"A1:{get_column_letter(len(header))}{last_row}"

wb.save('/workspace/ACH_Payments_2026-03-30.xlsx')
print("Saved!")
